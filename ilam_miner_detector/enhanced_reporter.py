"""
Enhanced Report Generation Module - PDF and Excel export.
Extends basic reporter with professional export formats.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from io import BytesIO

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available. Install with: pip install reportlab")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Install with: pip install openpyxl")

from .database import ScanRecord, HostRecord, get_db_manager
from .config_manager import get_config_manager

logger = logging.getLogger(__name__)


class EnhancedReportGenerator:
    """
    Enhanced report generator with PDF and Excel export capabilities.
    """
    
    def __init__(self):
        self.config = get_config_manager().get()
        self.db = get_db_manager()
        self.reports_dir = Path(self.config.reporting.reports_dir)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_timestamp(self) -> str:
        """Generate timestamp string for filenames."""
        return datetime.now().strftime(self.config.reporting.timestamp_format)
    
    def export_pdf(self, scan_id: int, output_path: Optional[str] = None) -> str:
        """
        Export scan results to PDF report.
        
        Args:
            scan_id: Scan ID to export
            output_path: Custom output path (optional)
            
        Returns:
            Path to exported PDF file
        """
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("ReportLab is required for PDF export. Install with: pip install reportlab")
        
        scan = self.db.get_scan(scan_id)
        if not scan:
            raise ValueError(f"Scan {scan_id} not found")
        
        hosts = self.db.get_hosts_by_scan(scan_id)
        miners = [h for h in hosts if h.is_miner_detected]
        
        # Determine output path
        if output_path is None:
            timestamp = self.generate_timestamp()
            output_path = self.reports_dir / f"scan_{scan_id}_{timestamp}.pdf"
        
        # Create PDF document
        doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Story elements
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#d32f2f'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#333333'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # Title
        story.append(Paragraph("🚨 Iranian Network Miner Detection Report", title_style))
        story.append(Spacer(1, 20))
        
        # Scan metadata
        metadata_data = [
            ["Scan Name:", scan.scan_name],
            ["CIDR Range:", scan.cidr_range],
            ["Start Time:", scan.start_time.strftime('%Y-%m-%d %H:%M:%S') if scan.start_time else 'N/A'],
            ["Status:", scan.status.title()],
            ["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 20))
        
        # Summary
        detection_rate = (scan.miners_detected / scan.total_hosts * 100) if scan.total_hosts > 0 else 0
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Hosts Scanned", str(scan.total_hosts)],
            ["Responsive Hosts", str(scan.responsive_hosts)],
            ["Miners Detected", str(scan.miners_detected)],
            ["Detection Rate", f"{detection_rate:.2f}%"],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Detected Miners
        if miners:
            story.append(Paragraph("Detected Miners", heading_style))
            
            miner_data = [["IP Address", "Type", "Confidence", "Open Ports"]]
            for miner in miners:
                import json
                try:
                    ports = json.loads(miner.open_ports) if miner.open_ports else []
                except:
                    ports = []
                
                miner_data.append([
                    miner.ip_address,
                    miner.miner_type or "Unknown",
                    f"{miner.confidence_score:.1f}%",
                    ", ".join(str(p) for p in ports[:5])
                ])
            
            miner_table = Table(miner_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 2*inch])
            miner_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5576c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffebee')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffebee')]),
            ]))
            story.append(miner_table)
            story.append(Spacer(1, 20))
        
        # All Hosts
        story.append(Paragraph("All Scanned Hosts", heading_style))
        
        host_data = [["IP Address", "Responsive", "Ping (ms)", "Miner", "Confidence"]]
        for host in hosts:
            host_data.append([
                host.ip_address,
                "Yes" if host.is_responsive else "No",
                f"{host.ping_time_ms:.2f}" if host.ping_time_ms else "N/A",
                "Yes" if host.is_miner_detected else "No",
                f"{host.confidence_score:.1f}%" if host.is_miner_detected else "-"
            ])
        
        host_table = Table(host_data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch], repeatRows=1)
        host_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        story.append(host_table)
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            "Generated by Iranian Network Miner Detection System",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)
        ))
        story.append(Paragraph(
            "This report is for authorized security auditing purposes only.",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)
        ))
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"PDF report saved to {output_path}")
        return str(output_path)
    
    def export_excel(self, scan_id: int, output_path: Optional[str] = None) -> str:
        """
        Export scan results to Excel workbook with multiple sheets.
        
        Args:
            scan_id: Scan ID to export
            output_path: Custom output path (optional)
            
        Returns:
            Path to exported Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        scan = self.db.get_scan(scan_id)
        if not scan:
            raise ValueError(f"Scan {scan_id} not found")
        
        hosts = self.db.get_hosts_by_scan(scan_id)
        miners = [h for h in hosts if h.is_miner_detected]
        
        # Determine output path
        if output_path is None:
            timestamp = self.generate_timestamp()
            output_path = self.reports_dir / f"scan_{scan_id}_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 1. Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._write_summary_sheet(ws_summary, scan, miners)
        
        # 2. Miners sheet
        ws_miners = wb.create_sheet("Detected Miners", 1)
        self._write_miners_sheet(ws_miners, miners)
        
        # 3. All Hosts sheet
        ws_hosts = wb.create_sheet("All Hosts", 2)
        self._write_hosts_sheet(ws_hosts, hosts)
        
        # 4. Analytics sheet
        ws_analytics = wb.create_sheet("Analytics", 3)
        self._write_analytics_sheet(ws_analytics, scan, miners, hosts)
        
        # Save workbook
        wb.save(str(output_path))
        
        logger.info(f"Excel report saved to {output_path}")
        return str(output_path)
    
    def _write_summary_sheet(self, ws, scan: ScanRecord, miners: List[HostRecord]):
        """Write summary sheet to Excel workbook."""
        # Title
        ws['A1'] = "Iranian Network Miner Detection Report"
        ws['A1'].font = Font(size=16, bold=True, color="AA0000")
        ws.merge_cells('A1:D1')
        
        # Metadata
        ws['A3'] = "Scan Metadata"
        ws['A3'].font = Font(size=12, bold=True)
        
        metadata_data = [
            ["Scan Name:", scan.scan_name],
            ["CIDR Range:", scan.cidr_range],
            ["Start Time:", scan.start_time.strftime('%Y-%m-%d %H:%M:%S') if scan.start_time else 'N/A'],
            ["End Time:", scan.end_time.strftime('%Y-%m-%d %H:%M:%S') if scan.end_time else 'N/A'],
            ["Status:", scan.status.title()],
        ]
        
        for i, (label, value) in enumerate(metadata_data, start=4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
        
        # Statistics
        ws['A10'] = "Statistics"
        ws['A10'].font = Font(size=12, bold=True)
        
        detection_rate = (scan.miners_detected / scan.total_hosts * 100) if scan.total_hosts > 0 else 0
        responsive_rate = (scan.responsive_hosts / scan.total_hosts * 100) if scan.total_hosts > 0 else 0
        
        stats_data = [
            ["Total Hosts Scanned", scan.total_hosts],
            ["Responsive Hosts", scan.responsive_hosts],
            ["Responsive Rate", f"{responsive_rate:.2f}%"],
            ["Miners Detected", scan.miners_detected],
            ["Detection Rate", f"{detection_rate:.2f}%"],
        ]
        
        for i, (label, value) in enumerate(stats_data, start=11):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _write_miners_sheet(self, ws, miners: List[HostRecord]):
        """Write detected miners sheet."""
        # Header
        headers = ["IP Address", "Miner Type", "Confidence Score", "Open Ports", "Banner Info"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        
        # Data rows
        import json
        for row_idx, miner in enumerate(miners, start=2):
            try:
                ports = json.loads(miner.open_ports) if miner.open_ports else []
                ports_str = ", ".join(str(p) for p in ports)
            except:
                ports_str = ""
            
            ws.cell(row=row_idx, column=1, value=miner.ip_address)
            ws.cell(row=row_idx, column=2, value=miner.miner_type or "Unknown")
            ws.cell(row=row_idx, column=3, value=f"{miner.confidence_score:.1f}%")
            ws.cell(row=row_idx, column=4, value=ports_str)
            ws.cell(row=row_idx, column=5, value=str(miner.banner_info)[:100])
            
            # Highlight row
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                )
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
    
    def _write_hosts_sheet(self, ws, hosts: List[HostRecord]):
        """Write all hosts sheet."""
        # Header
        headers = ["IP Address", "Responsive", "Ping Time (ms)", "Open Ports", 
                  "Miner Detected", "Miner Type", "Confidence"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
        # Data rows
        import json
        for row_idx, host in enumerate(hosts, start=2):
            try:
                ports = json.loads(host.open_ports) if host.open_ports else []
            except:
                ports = []
            
            ws.cell(row=row_idx, column=1, value=host.ip_address)
            ws.cell(row=row_idx, column=2, value="Yes" if host.is_responsive else "No")
            ws.cell(row=row_idx, column=3, value=f"{host.ping_time_ms:.2f}" if host.ping_time_ms else "")
            ws.cell(row=row_idx, column=4, value=", ".join(str(p) for p in ports[:5]))
            ws.cell(row=row_idx, column=5, value="Yes" if host.is_miner_detected else "No")
            ws.cell(row=row_idx, column=6, value=host.miner_type or "")
            ws.cell(row=row_idx, column=7, value=f"{host.confidence_score:.1f}%" if host.is_miner_detected else "")
            
            # Highlight miners
            if host.is_miner_detected:
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                    )
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
    
    def _write_analytics_sheet(self, ws, scan: ScanRecord, miners: List[HostRecord], hosts: List[HostRecord]):
        """Write analytics sheet."""
        # Title
        ws['A1'] = "Scan Analytics"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        # Miner types distribution
        ws['A3'] = "Miner Types Distribution"
        ws['A3'].font = Font(size=12, bold=True)
        
        miner_types = {}
        for miner in miners:
            miner_type = miner.miner_type or "Unknown"
            miner_types[miner_type] = miner_types.get(miner_type, 0) + 1
        
        row = 4
        for miner_type, count in sorted(miner_types.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=miner_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Confidence distribution
        row += 2
        ws[f'A{row}'] = "Confidence Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        confidence_ranges = {
            "High (80-100%)": 0,
            "Medium (50-79%)": 0,
            "Low (20-49%)": 0,
            "Very Low (0-19%)": 0
        }
        
        for miner in miners:
            score = miner.confidence_score
            if score >= 0.8:
                confidence_ranges["High (80-100%)"] += 1
            elif score >= 0.5:
                confidence_ranges["Medium (50-79%)"] += 1
            elif score >= 0.2:
                confidence_ranges["Low (20-49%)"] += 1
            else:
                confidence_ranges["Very Low (0-19%)"] += 1
        
        for range_name, count in confidence_ranges.items():
            ws.cell(row=row, column=1, value=range_name)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15


def get_enhanced_report_generator() -> EnhancedReportGenerator:
    """Get enhanced report generator instance."""
    return EnhancedReportGenerator()
